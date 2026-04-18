#!/usr/bin/env python3
"""
Hospital Price JSON Parser Bot
Converts hospital standard charges JSON into clean, searchable databases
Works with any hospital following CMS format
"""

import json
import requests
import csv
import os
from datetime import datetime
from pathlib import Path

class HospitalPriceParser:
    def __init__(self, json_url=None, json_file=None):
        """Initialize with either URL or local file"""
        self.data = None
        self.hospital_name = ""
        self.procedures = []
        self.output_dir = Path("C:\\Users\\Owner\\OneDrive\\Desktop\\Hospital_Pricing")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        if json_url:
            self._load_from_url(json_url)
        elif json_file:
            self._load_from_file(json_file)
    
    def _load_from_url(self, url):
        """Download and parse JSON from URL"""
        print(f"[*] Downloading from: {url}")
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            # Handle UTF-8 BOM
            self.data = json.loads(response.text.encode().decode('utf-8-sig'))
            print("[OK] Downloaded successfully")
        except Exception as e:
            print(f"[ERROR] Failed to download: {e}")
            return False
        return True
    
    def _load_from_file(self, file_path):
        """Load JSON from local file"""
        print(f"[*] Loading from file: {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                self.data = json.load(f)
            print("[OK] Loaded successfully")
        except Exception as e:
            print(f"[ERROR] Failed to load: {e}")
            return False
        return True
    
    def parse(self):
        """Extract hospital info and procedures"""
        if not self.data:
            print("[ERROR] No data loaded")
            return False
        
        # Get hospital info
        self.hospital_name = self.data.get("hospital_name", "Unknown Hospital")
        address = self.data.get("hospital_address", ["Unknown"])[0]
        last_updated = self.data.get("last_updated_on", "Unknown")
        
        print(f"\n[HOSPITAL INFO]")
        print(f"  Name: {self.hospital_name}")
        print(f"  Address: {address}")
        print(f"  Last Updated: {last_updated}")
        
        # Extract procedures
        print(f"\n[EXTRACTING PROCEDURES]")
        for charge in self.data.get("standard_charge_information", []):
            description = charge.get("description", "").strip()
            if not description:
                continue
            
            # Get all codes
            codes = charge.get("code_information", [])
            cpt_code = ""
            rc_code = ""
            hcpcs_code = ""
            cdm_code = ""
            
            for code_info in codes:
                code_type = code_info.get("type", "")
                code_value = code_info.get("code", "")
                
                if code_type == "CPT":
                    cpt_code = code_value
                elif code_type == "RC":
                    rc_code = code_value
                elif code_type == "HCPCS":
                    hcpcs_code = code_value
                elif code_type == "CDM":
                    cdm_code = code_value
            
            # Get pricing for each charge detail
            for charge_detail in charge.get("standard_charges", []):
                gross_charge = charge_detail.get("gross_charge", "N/A")
                discounted_cash = charge_detail.get("discounted_cash", "N/A")
                billing_class = charge_detail.get("billing_class", "")
                setting = charge_detail.get("setting", "")
                
                # Format prices
                try:
                    gross_formatted = "${:,.2f}".format(float(gross_charge))
                    discount_formatted = "${:,.2f}".format(float(discounted_cash))
                    savings = "${:,.2f}".format(float(gross_charge) - float(discounted_cash))
                except (ValueError, TypeError):
                    gross_formatted = str(gross_charge)
                    discount_formatted = str(discounted_cash)
                    savings = "N/A"
                
                # Categorize procedure
                category = self._categorize_procedure(description, rc_code, cpt_code)
                
                procedure = {
                    "Procedure Name": description,
                    "Category": category,
                    "CPT Code": cpt_code,
                    "RC Code": rc_code,
                    "HCPCS Code": hcpcs_code,
                    "CDM Code": cdm_code,
                    "Billing Class": billing_class,
                    "Setting": setting,
                    "Gross Charge": gross_formatted,
                    "Discounted Cash": discount_formatted,
                    "Savings": savings,
                    "Raw Gross": str(gross_charge),
                    "Raw Discount": str(discounted_cash)
                }
                
                self.procedures.append(procedure)
        
        # Remove duplicates and sort
        self.procedures = list({v['Procedure Name']: v for v in self.procedures}.values())
        self.procedures.sort(key=lambda x: x['Procedure Name'].lower())
        
        print(f"[OK] Extracted {len(self.procedures)} unique procedures")
        return True
    
    def _categorize_procedure(self, name, rc_code, cpt_code):
        """Categorize procedure by keywords and codes"""
        name_lower = name.lower()
        rc_code = str(rc_code).lower()
        
        # Categorization rules
        if any(x in name_lower for x in ['surgical', 'surgery', 'operation', 'incision', 'ablation', 'removal', 'excision']):
            return "Surgical"
        elif any(x in name_lower for x in ['mri', 'ct scan', 'x-ray', 'ultrasound', 'imaging', 'radiograph', 'echocardiogram', 'ekg', 'echo']):
            return "Imaging"
        elif any(x in name_lower for x in ['blood test', 'lab', 'panel', 'culture', 'urinalysis', 'pathology']):
            return "Laboratory"
        elif any(x in name_lower for x in ['drug', 'medication', 'injection', 'infusion', 'vaccine', 'serum']):
            return "Pharmacy"
        elif any(x in name_lower for x in ['room', 'bed', 'icu', 'er', 'emergency', 'ward', 'nursery']):
            return "Room & Board"
        elif any(x in name_lower for x in ['therapy', 'rehab', 'physical', 'occupational', 'speech']):
            return "Rehabilitation"
        elif any(x in name_lower for x in ['anesthesia', 'supplies', 'equipment', 'device']):
            return "Supplies & Equipment"
        elif rc_code.startswith('25'):
            return "Pharmacy"
        elif rc_code.startswith('27'):
            return "Supplies & Equipment"
        elif rc_code in ['111', '112', '114', '117', '121', '122', '124', '128', '170', '200', '206', '210']:
            return "Room & Board"
        else:
            return "Other Services"
    
    def export_csv(self):
        """Export to CSV"""
        filename = f"{self.hospital_name.replace(' ', '_')}_Procedures_{datetime.now().strftime('%Y%m%d')}.csv"
        filepath = self.output_dir / filename
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                fieldnames = [
                    "Procedure Name",
                    "Category",
                    "CPT Code",
                    "RC Code",
                    "HCPCS Code",
                    "CDM Code",
                    "Billing Class",
                    "Setting",
                    "Gross Charge",
                    "Discounted Cash",
                    "Savings"
                ]
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for proc in self.procedures:
                    row = {k: proc[k] for k in fieldnames}
                    writer.writerow(row)
            
            print(f"[OK] Exported CSV: {filepath}")
            return str(filepath)
        except Exception as e:
            print(f"[ERROR] Failed to export CSV: {e}")
            return None
    
    def export_excel(self):
        """Export to Excel with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("[NOTE] openpyxl not installed. Install with: pip install openpyxl")
            return None
        
        filename = f"{self.hospital_name.replace(' ', '_')}_Procedures_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = self.output_dir / filename
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Procedures"
            
            # Headers
            headers = [
                "Procedure Name",
                "Category",
                "CPT Code",
                "RC Code",
                "HCPCS Code",
                "CDM Code",
                "Billing Class",
                "Setting",
                "Gross Charge",
                "Discounted Cash",
                "Savings"
            ]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Add data rows
            for proc in self.procedures:
                ws.append([
                    proc["Procedure Name"],
                    proc["Category"],
                    proc["CPT Code"],
                    proc["RC Code"],
                    proc["HCPCS Code"],
                    proc["CDM Code"],
                    proc["Billing Class"],
                    proc["Setting"],
                    proc["Gross Charge"],
                    proc["Discounted Cash"],
                    proc["Savings"]
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 14
            ws.column_dimensions['J'].width = 16
            ws.column_dimensions['K'].width = 14
            
            wb.save(filepath)
            print(f"[OK] Exported Excel: {filepath}")
            return str(filepath)
        except Exception as e:
            print(f"[ERROR] Failed to export Excel: {e}")
            return None
    
    def export_json(self):
        """Export to JSON for web database"""
        filename = f"{self.hospital_name.replace(' ', '_')}_Procedures_{datetime.now().strftime('%Y%m%d')}.json"
        filepath = self.output_dir / filename
        
        try:
            # Remove raw values for export
            clean_procedures = []
            for proc in self.procedures:
                clean_proc = {k: v for k, v in proc.items() if not k.startswith('Raw')}
                clean_procedures.append(clean_proc)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump({
                    "hospital": self.hospital_name,
                    "exported": datetime.now().isoformat(),
                    "total_procedures": len(clean_procedures),
                    "procedures": clean_procedures
                }, f, indent=2)
            
            print(f"[OK] Exported JSON: {filepath}")
            return str(filepath)
        except Exception as e:
            print(f"[ERROR] Failed to export JSON: {e}")
            return None
    
    def export_by_category(self):
        """Export separate files by category"""
        categories = {}
        for proc in self.procedures:
            cat = proc['Category']
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(proc)
        
        for category, procs in categories.items():
            filename = f"{self.hospital_name.replace(' ', '_')}_{category.replace(' & ', '_').replace(' ', '_')}__{datetime.now().strftime('%Y%m%d')}.csv"
            filepath = self.output_dir / filename
            
            try:
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    fieldnames = [
                        "Procedure Name",
                        "CPT Code",
                        "RC Code",
                        "HCPCS Code",
                        "Billing Class",
                        "Gross Charge",
                        "Discounted Cash",
                        "Savings"
                    ]
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    for proc in procs:
                        row = {k: proc.get(k, "") for k in fieldnames}
                        writer.writerow(row)
                
                print(f"  [{len(procs)}] {category} -> {filepath.name}")
            except Exception as e:
                print(f"  [ERROR] Failed to export {category}: {e}")


def main():
    """Main execution"""
    print("\n" + "="*70)
    print("HOSPITAL PRICE JSON PARSER BOT")
    print("="*70 + "\n")
    
    # Example: Parse Riverside Methodist
    url = "https://www.ohiohealth.com/siteassets/patients-and-visitors/preparing-for-your-visit/out-of-pocket-estimates/314394942_riverside-methodist-hospital_standardcharges.json"
    
    parser = HospitalPriceParser(json_url=url)
    
    if parser.parse():
        print("\n[EXPORTING]")
        csv_file = parser.export_csv()
        xlsx_file = parser.export_excel()
        json_file = parser.export_json()
        
        print("\n[BY CATEGORY]")
        parser.export_by_category()
        
        print("\n" + "="*70)
        print("COMPLETE!")
        print("="*70)
        print(f"\nHospital: {parser.hospital_name}")
        print(f"Procedures: {len(parser.procedures)}")
        print(f"Output Directory: {parser.output_dir}")
        print("\nFiles created:")
        if csv_file:
            print(f"  • CSV: {Path(csv_file).name}")
        if xlsx_file:
            print(f"  • Excel: {Path(xlsx_file).name}")
        if json_file:
            print(f"  • JSON: {Path(json_file).name}")


if __name__ == "__main__":
    main()
