#!/usr/bin/env python3
"""
Riverside Methodist Hospital Price Listing Extractor
Converts complex JSON hospital pricing into clean, searchable spreadsheet
"""

import json
import requests
import csv
from datetime import datetime
import sys

# Download the JSON file
print("Downloading hospital pricing data from Ohio Health...")
url = "https://www.ohiohealth.com/siteassets/patients-and-visitors/preparing-for-your-visit/out-of-pocket-estimates/314394942_riverside-methodist-hospital_standardcharges.json"

try:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    # Handle UTF-8 BOM encoding
    data = json.loads(response.text.encode().decode('utf-8-sig'))
    print("[OK] Downloaded hospital pricing data")
except Exception as e:
    print(f"[ERROR] Error downloading: {e}")
    sys.exit(1)

# Extract hospital info
hospital_name = data.get("hospital_name", "Unknown Hospital")
hospital_address = data.get("hospital_address", ["Unknown"])[0]
last_updated = data.get("last_updated_on", "Unknown")

print("\nHospital: " + hospital_name)
print("Address: " + hospital_address)
print("Last Updated: " + last_updated)

# Extract pricing information
pricing_rows = []

for charge in data.get("standard_charge_information", []):
    description = charge.get("description", "")
    
    # Get codes
    codes = charge.get("code_information", [])
    cpt_code = ""
    rc_code = ""
    hcpcs_code = ""
    
    for code_info in codes:
        code_type = code_info.get("type", "")
        code_value = code_info.get("code", "")
        
        if code_type == "CPT":
            cpt_code = code_value
        elif code_type == "RC":
            rc_code = code_value
        elif code_type == "HCPCS":
            hcpcs_code = code_value
    
    # Get pricing
    for charge_detail in charge.get("standard_charges", []):
        gross_charge = charge_detail.get("gross_charge", "N/A")
        discounted_cash = charge_detail.get("discounted_cash", "N/A")
        billing_class = charge_detail.get("billing_class", "")
        
        # Format prices
        try:
            gross_formatted = "${:,.2f}".format(float(gross_charge)) if isinstance(gross_charge, (int, float)) else str(gross_charge)
            discount_formatted = "${:,.2f}".format(float(discounted_cash)) if isinstance(discounted_cash, (int, float)) else str(discounted_cash)
        except:
            gross_formatted = str(gross_charge)
            discount_formatted = str(discounted_cash)
        
        try:
            savings = "${:,.2f}".format(float(gross_charge) - float(discounted_cash)) if isinstance(gross_charge, (int, float)) and isinstance(discounted_cash, (int, float)) else "N/A"
        except:
            savings = "N/A"
        
        row = {
            "Procedure/Test Name": description.strip(),
            "CPT Code": cpt_code,
            "RC Code": rc_code,
            "HCPCS Code": hcpcs_code,
            "Billing Class": billing_class,
            "Gross Charge": gross_formatted,
            "Discounted Cash Price": discount_formatted,
            "Savings": savings
        }
        
        pricing_rows.append(row)

# Sort alphabetically by procedure name
pricing_rows.sort(key=lambda x: x["Procedure/Test Name"].lower())

print("\n[OK] Extracted {} procedures/tests".format(len(pricing_rows)))

# Create CSV file
csv_filename = f"Riverside_Methodist_Hospital_Pricing_{datetime.now().strftime('%Y%m%d')}.csv"
csv_path = f"C:\\Users\\Owner\\OneDrive\\Desktop\\{csv_filename}"

try:
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            "Procedure/Test Name",
            "CPT Code",
            "RC Code",
            "HCPCS Code",
            "Billing Class",
            "Gross Charge",
            "Discounted Cash Price",
            "Savings"
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(pricing_rows)
    
    print("\n[OK] Created CSV: " + csv_path)
except Exception as e:
    print("\n[ERROR] Error creating CSV: " + str(e))
    sys.exit(1)

# Also create Excel file with formatting
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hospital Pricing"
    
    # Add header
    headers = [
        "Procedure/Test Name",
        "CPT Code",
        "RC Code",
        "HCPCS Code",
        "Billing Class",
        "Gross Charge",
        "Discounted Cash Price",
        "Savings"
    ]
    
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_data in pricing_rows:
        ws.append([
            row_data["Procedure/Test Name"],
            row_data["CPT Code"],
            row_data["RC Code"],
            row_data["HCPCS Code"],
            row_data["Billing Class"],
            row_data["Gross Charge"],
            row_data["Discounted Cash Price"],
            row_data["Savings"]
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    
    excel_filename = "Riverside_Methodist_Hospital_Pricing_{}.xlsx".format(datetime.now().strftime('%Y%m%d'))
    excel_path = "C:\\Users\\Owner\\OneDrive\\Desktop\\" + excel_filename
    
    wb.save(excel_path)
    print("[OK] Created Excel: " + excel_path)
    
except ImportError:
    print("\n[NOTE] openpyxl not installed for Excel formatting")
    print("   Install with: pip install openpyxl")

print("\n" + "="*60)
print("COMPLETE!")
print("="*60)
print("\nFiles created in: C:\\Users\\Owner\\OneDrive\\Desktop\\")
print("  * " + csv_filename)
if 'excel_path' in locals():
    print("  * " + excel_filename)
print("\n{} procedures extracted and sorted alphabetically".format(len(pricing_rows)))
print("\nYou can now:")
print("  1. Open in Excel/LibreOffice")
print("  2. Search by procedure name")
print("  3. Sort by price")
print("  4. Filter by code type")
print("  5. Copy to your website")
